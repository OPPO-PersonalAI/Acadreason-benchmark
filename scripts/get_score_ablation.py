#!/usr/bin/env python
# coding=utf-8
# Copyright 2025 The OPPO Inc. PersonalAI team. All rights reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
"""
Comprehensive experiment score statistics script
Process both main experiment and ablation experiment data, generate Excel file with multiple sheets
python scripts/get_score_ablation.py results/judge -o my_results.xlsx -v -c 10
Usage:
1. Basic usage:
   python scripts/get_score_ablation.py results/judge
   
2. Specify output file:
   python scripts/get_score_ablation.py results/judge -o my_results.xlsx
   
3. Show verbose output:
   python scripts/get_score_ablation.py results/judge -v

4. Custom expected data count:
   python scripts/get_score_ablation.py results/judge -c 50

5. Complete usage:
   python scripts/get_score_ablation.py results/judge -o my_results.xlsx -v -c 50

Input data structure (fixed definition):
results/judge/
├── judge_infer_50_hints0/     # Ablation experiment: no hint
├── judge_infer_50_hints1/     # Ablation experiment: Hint1
├── judge_infer_50_hints2/     # Ablation experiment: Hint2
├── judge_infer_50_hints3/     # Ablation experiment: Hint3
├── judge_infer_50_hints4/     # Ablation experiment: Hint1+Hint2+Hint3


Output format:
- Excel file contains multiple sheets:
  1. "main": Main experiment results (classified by category, format: aspect1%/aspect2%)
  2. One sheet per hints condition (same format as main, with discipline columns)
  3. "ablation experiment": Ablation experiment results (specified models, aspect1%/aspect2% format for 5 hint conditions)

Automatic model support (no manual configuration needed):
- Automatically extract model names from filenames
- Intelligently generate display names
- Support complex model name formats
- Maintain consistent sorting rules

Calculation method:
1. aspect1 total score = sum of numerator scores / data count (each question max 1 point, only numerator counted)
2. aspect2 total score = score / sum(num_checklist)  
3. total score = (aspect1 numerator sum + aspect2 raw total) / (total checklist count + data count)
4. Calculate sum by all available categories

Data completeness check:
- Each category expects 50 data entries by default (customizable via -c parameter)
- Not displayed during normal data processing, only shown at the end for problematic data
- Missing data automatically saved to timestamped txt file
"""

import json
import argparse
import os
import glob
import re
from typing import Dict, List, Tuple, Set
from collections import defaultdict
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime

# Global variable to collect all completeness issues
all_completeness_issues = []


# ========== Model Management System ==========

class ModelManager:
    """Model configuration and name manager"""
    
    # Predefined model configurations (maintain backward compatibility and sorting priority)
    PREDEFINED_MODELS = {
        # Core models (high priority)
        'gpt41': {'display_name': 'gpt4.1', 'priority': 100},
        'gpt4o': {'display_name': 'gpt4o', 'priority': 99},
        'gpt5': {'display_name': 'gpt5', 'priority': 98}, 
        'gpt5mini': {'display_name': 'gpt5mini', 'priority': 97},
        'claude4': {'display_name': 'claude4', 'priority': 96},
        'o1': {'display_name': 'o1', 'priority': 95},
        'o3': {'display_name': 'o3', 'priority': 94},
        'o3-dr': {'display_name': 'o3-dr', 'priority': 93},
        'o4_mini_dr': {'display_name': 'o4_mini_dr', 'priority': 92},
        
        # DeepSeek series
        'deepseekv3': {'display_name': 'deepseekv3', 'priority': 85},
        'deepseekv31': {'display_name': 'deepseekv31', 'priority': 84},
        'deepseekr1': {'display_name': 'deepseekr1', 'priority': 83},
        
        # Gemini series 
        'gemini25pro': {'display_name': 'gemini2.5pro', 'priority': 75},
        'gemini_2.5_pro_deepsearch_async': {'display_name': 'gemini2.5pro_deepsearch', 'priority': 74},
        'gemini_2.5_flash_deepsearch_async': {'display_name': 'gemini2.5flash_deepsearch', 'priority': 73},
        
        # AFM series
        'AFM_sft': {'display_name': 'AFM_sft', 'priority': 70},
        'AFM_rl': {'display_name': 'AFM_rl', 'priority': 69},
        
        # Other models
        'oagent': {'display_name': 'oagent', 'priority': 65},
        'gptoss': {'display_name': 'gptoss', 'priority': 64},
        'kimik2': {'display_name': 'kimik2', 'priority': 63},
        'qwen3': {'display_name': 'qwen3', 'priority': 62},
        'seedoss': {'display_name': 'seedoss', 'priority': 61},
    }
    
    def __init__(self):
        self.discovered_models = {}  # New models discovered at runtime
    
    def extract_model_name(self, filename: str) -> str:
        """
        Intelligently extract model name, support various filename formats
        """
        if not filename or not isinstance(filename, str):
            return filename
        
        # Remove file extension
        base_name = filename
        if base_name.endswith('.jsonl'):
            base_name = base_name[:-6]
        
        # Handle different filename formats
        if base_name.startswith('judged_'):
            # Format: judged_{model}_{judge}.jsonl
            name_part = base_name[7:]  # Remove 'judged_' prefix
            
            # Filename format: judged_{model_being_judged}_{judge_model}.jsonl
            # We need to identify the model being judged, not the judge model
            
            # Known judge model list (these should be excluded, not treated as models being judged)
            judge_models = {
                'gpt5mini', 'gpt41', 'gpt4o', 'gpt5', 'claude4', 'o1', 'o3'
            }
            
            # Special handling for complex model names (models being judged, match by length priority)
            complex_patterns = [
                'gemini_2.5_pro_deepsearch_async',
                'gemini_2.5_flash_deepsearch_async', 
                'deepseekv31',
                'deepseekv3',
                'deepseekr1',
                'o4_mini_dr',
                'o3-dr',
                'gemini25pro',
                'AFM_sft',
                'AFM_rl',
                'oagent',
                'gptoss',
                'kimik2',
                'qwen3',
                'seedoss'
            ]
            
            # Sort by length in descending order, ensure long patterns match first
            complex_patterns.sort(key=len, reverse=True)
            
            # First try to match model being judged (exclude judge models)
            for pattern in complex_patterns:
                if pattern in name_part:
                    # Ensure this is not part of judge model
                    remaining = name_part.replace(pattern, '')
                    # Check if remaining part is a known judge model
                    remaining_parts = [p for p in remaining.split('_') if p]
                    if any(part in judge_models for part in remaining_parts):
                        return pattern
            
            # Generic parsing: split and intelligently identify model being judged
            parts = name_part.split('_')
            if len(parts) >= 2:
                # Try to identify if last part is judge model
                last_part = parts[-1]
                if last_part in judge_models:
                    # Last part is judge model, previous part is model being judged
                    model_name = '_'.join(parts[:-1])
                    return model_name
                else:
                    # Cannot determine, use original logic
                    model_name = '_'.join(parts[:-1])
                    return model_name
            elif len(parts) == 1:
                return parts[0]
                
        elif base_name.startswith('judge_test_'):
            # 格式: judge_test_first_{model}_{judge}.jsonl
            name_part = base_name[11:]  # 移除 'judge_test_' 前缀
            
            # Remove common suffix patterns
            for suffix in ['_gpt41', '_gpt5mini']:
                if name_part.endswith(suffix):
                    name_part = name_part[:-len(suffix)]
                    break
            
            # Find known model names in remaining part
            for pattern in ['gpt41', 'gpt5', 'claude4', 'deepseekv3', 'deepseekr1', 
                           'o1', 'o3', 'gemini25pro', 'o4_mini_dr', 'AFM_sft', 'AFM_rl']:
                if pattern in name_part:
                    return pattern
        
        # If none of the above match, return processed base name
        return base_name
    
    def get_display_name(self, model_name: str) -> str:
        """
        获取Model的显示名称
        """
        if not model_name:
            return model_name
            
        # First check predefined models
        if model_name in self.PREDEFINED_MODELS:
            return self.PREDEFINED_MODELS[model_name]['display_name']
        
        # Check already discovered models
        if model_name in self.discovered_models:
            return self.discovered_models[model_name]['display_name']
        
        # Generate display name for new model
        display_name = self._generate_display_name(model_name)
        
        # Record newly discovered model (medium priority)
        self.discovered_models[model_name] = {
            'display_name': display_name,
            'priority': 50  # New models default to medium priority
        }
        
        print(f"🆕 Discovered new model: {model_name} -> {display_name}")
        
        return display_name
    
    def _generate_display_name(self, model_name: str) -> str:
        """
        Intelligently generate display name for new model
        """
        # Simple cleaning and formatting
        display_name = model_name.replace('_', '.')
        
        # Handle common patterns
        patterns = [
            (r'gpt(\d+)', r'gpt\1'),
            (r'claude(\d+)', r'claude\1'),  
            (r'o(\d+)', r'o\1'),
            (r'deepseek(v\d+)', r'deepseek\1'),
            (r'gemini(\d+\.?\d*)', r'gemini\1'),
        ]
        
        for pattern, replacement in patterns:
            display_name = re.sub(pattern, replacement, display_name, flags=re.IGNORECASE)
        
        return display_name
    
    def get_priority(self, model_name: str) -> int:
        """
        Get model's sorting priority
        """
        if model_name in self.PREDEFINED_MODELS:
            return self.PREDEFINED_MODELS[model_name]['priority']
        elif model_name in self.discovered_models:
            return self.discovered_models[model_name]['priority']
        else:
            # Unknown models default to lowest priority
            return 0
    
    def get_sorted_models(self, available_models: Set[str]) -> List[Tuple[str, str]]:
        """
        Get sorted model list: (model_name, display_name)
        """
        # Ensure all models have display names
        for model in available_models:
            self.get_display_name(model)  # This will automatically register new models
        
        # Sort by priority, then by name if priority is the same
        sorted_models = sorted(
            available_models,
            key=lambda x: (-self.get_priority(x), x)
        )
        
        return [(model, self.get_display_name(model)) for model in sorted_models]


# Global model manager instance
model_manager = ModelManager()


# ========== Bench query -> num_checklist mapping ==========
_GLOBAL_BENCH_MAP = None
_GLOBAL_BENCH_DATA = None  # Store complete bench data for field matching


def _normalize_query(q: str) -> str:
    if not q:
        return ""
    q = ' '.join(q.split())
    q = q.replace(' | ', ' ').replace('|', ' ')
    return ' '.join(q.split()).strip()


def _load_bench_map(default_path: str = 'data/raw/bench_50.jsonl') -> Dict[str, int]:
    bench_map: Dict[str, int] = {}
    try:
        if not os.path.exists(default_path):
            return bench_map
        with open(default_path, 'r', encoding='utf-8') as f:
            for line in f:
                if not line.strip():
                    continue
                try:
                    obj = json.loads(line)
                except Exception:
                    continue
                q = obj.get('query') or obj.get('original_query') or ''
                ncl = obj.get('num_checklist')
                if isinstance(ncl, int) and ncl > 0:
                    bench_map[q] = ncl
                    bench_map[_normalize_query(q)] = ncl
    except Exception:
        pass
    return bench_map


def _get_bench_map() -> Dict[str, int]:
    global _GLOBAL_BENCH_MAP
    if _GLOBAL_BENCH_MAP is None:
        _GLOBAL_BENCH_MAP = _load_bench_map()
    return _GLOBAL_BENCH_MAP


def _load_bench_data(default_path: str = 'data/raw/bench_50.jsonl') -> Dict[str, dict]:
    """Load complete bench data for field matching"""
    bench_data: Dict[str, dict] = {}
    try:
        if not os.path.exists(default_path):
            return bench_data
        with open(default_path, 'r', encoding='utf-8') as f:
            for line in f:
                if not line.strip():
                    continue
                try:
                    obj = json.loads(line)
                    query = obj.get('query', '').strip()
                    if query:
                        # Use original query as key for exact matching
                        bench_data[query] = obj
                except Exception:
                    continue
    except Exception:
        pass
    return bench_data


def _get_bench_data() -> Dict[str, dict]:
    global _GLOBAL_BENCH_DATA
    if _GLOBAL_BENCH_DATA is None:
        _GLOBAL_BENCH_DATA = _load_bench_data()
    return _GLOBAL_BENCH_DATA


def get_category_from_bench(query: str) -> str:
    """Get category from bench_50.jsonl by matching query"""
    bench_data = _get_bench_data()
    if query in bench_data:
        return bench_data[query].get('category', '')
    return ''


def smart_categorize(data: dict) -> str:
    """
    When category is null, first try to match from bench_50.jsonl, then intelligently identify category
    
    Args:
        data: Data dictionary containing query and other fields
        
    Returns:
        Identified category name, returns 'Unknown' if unable to identify
    """
    # First try to match from bench_50.jsonl
    query = data.get('query', '')
    if query:
        bench_category = get_category_from_bench(query.strip())
        if bench_category:
            return bench_category
    
    # If unable to match from bench_50.jsonl, use intelligent identification
    query_lower = query.lower()
    sheet_name = data.get('sheet_name', '').lower()
    
    # Math keywords
    math_keywords = [
        'hilbert', 'samuel', 'multiplicity', 'cohen', 'macaulay', 'rings', 
        'characteristic', 'frobenius', 'gorenstein', 'algebra', 'algebraic',
        'theorem', 'lemma', 'proof', 'mathematical', 'equation', 'formula',
        'function', 'matrix', 'polynomial', 'topology', 'geometry', 'calculus',
        'analysis', 'number theory', 'combinatorics', 'permutation', 'bijection',
        'motzkin', 'fibonacci', 'probability', 'stochastic', 'quantum', 'optimization'
    ]
    
    # Computer Science keywords
    cs_keywords = [
        'algorithm', 'data structure', 'programming', 'software', 'computer',
        'machine learning', 'artificial intelligence', 'neural network',
        'database', 'network', 'security', 'cryptography', 'blockchain'
    ]
    
    # Philosophy keywords
    philosophy_keywords = [
        'philosophy', 'philosophical', 'ethics', 'moral', 'ontology',
        'epistemology', 'metaphysics', 'logic', 'phenomenology'
    ]
    
    # Law keywords
    law_keywords = [
        'law', 'legal', 'court', 'justice', 'rights', 'constitution',
        'legislation', 'jurisprudence', 'contract', 'tort'
    ]
    
    # Economics keywords
    economics_keywords = [
        'economic', 'economics', 'market', 'trade', 'finance', 'fiscal',
        'monetary', 'GDP', 'inflation', 'supply', 'demand'
    ]
    
    # Check keywords for each category
    if any(keyword in query_lower for keyword in math_keywords):
        return 'Math'
    elif any(keyword in query_lower for keyword in cs_keywords):
        return 'Computer Science'
    elif any(keyword in query_lower for keyword in philosophy_keywords):
        return 'philosophy'
    elif any(keyword in query_lower for keyword in law_keywords):
        return 'Law'
    elif any(keyword in query_lower for keyword in economics_keywords):
        return 'economics'
    
    # If sheet_name contains discipline information, can also be used as reference
    if 'math' in sheet_name:
        return 'Math'
    elif 'computer' in sheet_name or 'cs' in sheet_name:
        return 'Computer Science'
    elif 'philosophy' in sheet_name:
        return 'philosophy'
    elif 'law' in sheet_name:
        return 'Law'
    elif 'econ' in sheet_name:
        return 'economics'
    
    return 'Unknown'


def parse_score_ratio(score_str: str) -> Tuple[float, float]:
    """
    Parse score ratio string, e.g. "1/2" -> (1.0, 2.0)
    
    Args:
        score_str: Score string, format "numerator/denominator"
        
    Returns:
        Tuple of (numerator, denominator)
    """
    try:
        if '/' in score_str:
            numerator, denominator = score_str.split('/')
            return float(numerator), float(denominator)
        else:
            # If no denominator, default denominator is 1
            return float(score_str), 1.0
    except (ValueError, AttributeError):
        return 0.0, 1.0




def calculate_scores_for_file(jsonl_file: str) -> Dict:
    """
    Calculate score statistics for all data in a single jsonl file, classified by category
    
    Args:
        jsonl_file: Input jsonl file path
        
    Returns:
        Dictionary containing classification calculation results
    """
    if not os.path.exists(jsonl_file):
        raise FileNotFoundError(f"File does not exist: {jsonl_file}")
    
    # Category statistics - use defaultdict to automatically create nested dict
    category_stats = defaultdict(lambda: {
        'count': 0,
        'aspect1_score': 0.0,
        'aspect1_denominator': 0.0,
        'aspect2_score': 0.0,
        'num_checklist': 0
    })
    
    print(f"Processing file: {jsonl_file}")
    
    with open(jsonl_file, 'r', encoding='utf-8') as f:
        for line_num, line in enumerate(f, 1):
            try:
                data = json.loads(line.strip())
                
                # Check if scores field exists
                if 'scores' not in data:
                    print(f"Warning: file{jsonl_file}第{line_num}line missing scores field, skipping")
                    continue
                
                scores = data['scores']
                
                # Parse aspect1 score (convert multi-point scale to 1-point: numerator>0 counts as 1, else 0)
                aspect1_str = scores.get('aspect1', '0/1')
                aspect1_score, aspect1_max = parse_score_ratio(aspect1_str)
                # Convert to pass/fail
                aspect1_raw_score = 1.0 if aspect1_score > 0 else 0.0
                aspect1_max = 1.0
                
                # Parse aspect2 score
                aspect2_str = scores.get('aspect2', '0/1')
                aspect2_score, aspect2_max = parse_score_ratio(aspect2_str)
                
                # Get num_checklist: prioritize scores.num_checklist;
                # if missing, match from bench_50.jsonl by query; fallback to aspect2 denominator if still missing
                num_checklist = scores.get('num_checklist', None)
                if not isinstance(num_checklist, (int, float)) or num_checklist <= 0:
                    # bench lookup
                    bench_map = _get_bench_map()
                    q = data.get('query') or data.get('original_query') or ''
                    num_checklist = bench_map.get(q) or bench_map.get(_normalize_query(q))
                if not isinstance(num_checklist, (int, float)) or num_checklist <= 0:
                    num_checklist = aspect2_max
                
                # Get category (from raw data)
                category = data.get('category', 'Unknown')
                # If category is None or empty, first try to match from bench, then intelligently identify category
                if category is None or category == '':
                    # Try to match and get complete field info from bench_50.jsonl
                    query = data.get('query', '')
                    if query:
                        bench_data = _get_bench_data()
                        matched_data = bench_data.get(query.strip())
                        if matched_data:
                            # Successfully matched bench data, use its category
                            category = matched_data.get('category', 'Unknown')
                            print(f"✅ Matched and supplemented category from bench_50.jsonl: {category}，query: {query[:50]}...")
                        else:
                            # Unable to match, use intelligent identification
                            category = smart_categorize(data)
                            if category == 'Unknown':
                                print(f"Warning: file{jsonl_file}第{line_num}line category is null and cannot be auto-identified, set to Unknown")
                                print(f"      Content preview: {data.get('query', '')[:100]}...")
                    else:
                        category = 'Unknown'
                
                # Category statistics
                category_stats[category]['count'] += 1
                category_stats[category]['aspect1_score'] += aspect1_raw_score
                category_stats[category]['aspect1_denominator'] += aspect1_max
                category_stats[category]['aspect2_score'] += aspect2_score
                category_stats[category]['num_checklist'] += num_checklist
                
            except json.JSONDecodeError as e:
                print(f"Warning: file{jsonl_file}第{line_num}line JSON parsing error, skipping: {e}")
                continue
            except Exception as e:
                print(f"Warning: file{jsonl_file}第{line_num}line processing error, skipping: {e}")
                continue
    
    # Calculate category results
    category_results = {}
    for category, stats in category_stats.items():
        if stats['count'] > 0:  # Only calculate categories with data
            category_results[category] = calculate_category_scores(
                stats['count'], 
                stats['aspect1_score'], 
                stats['aspect1_denominator'],
                stats['aspect2_score'], 
                stats['num_checklist']
            )
    
    return category_results


def calculate_category_scores(data_count: int, aspect1_score: float, aspect1_denominator: float, aspect2_score: float, num_checklist: int) -> Dict:
    """
    Calculate scores for a single category
    
    Args:
        data_count: data count
        aspect1_score: aspect1 numerator cumulative score (not normalized, only take numerator)
        aspect1_denominator: aspect1 denominator cumulative (usually number of questions)
        aspect2_score: aspect2 total score
        num_checklist: total checklist count
        
    Returns:
        calculation result dictionary
    """
    # aspect1 total score = sum of numerator scores / number of questions (each question max 1 point)
    aspect1_total = aspect1_score / aspect1_denominator if aspect1_denominator > 0 else 0

    # aspect2 total score = 得分/sum(num_checklist)
    aspect2_total = aspect2_score / num_checklist if num_checklist > 0 else 0
    
    # total score = (aspect1 raw total + aspect2 raw total) / (total checklist count + aspect1 denominator cumulative)
    denominator = num_checklist + aspect1_denominator
    final_total = (aspect1_score + aspect2_score) / denominator if denominator > 0 else 0
    
    return {
        'data_count': data_count,
        'num_checklist': num_checklist,
        'raw_scores': {
            'aspect1_score': aspect1_score,
            'aspect1_denominator': aspect1_denominator,
            'aspect2_score': aspect2_score
        },
        'calculated_scores': {
            'aspect1_total': aspect1_total,
            'aspect2_total': aspect2_total,
            'final_total': final_total
        },
        'formatted_score': f"{aspect1_total*100:.1f}/{aspect2_total*100:.1f}"
    }


def check_data_completeness(category_results: Dict, expected_count_per_category: int = 50) -> Dict:
    """
    Check data completeness，确保每个category都有预期的data count
    
    Args:
        category_results: Calculation results for each category
        expected_count_per_category: 每个category预期的data count
        
    Returns:
        Dictionary containing completeness check results
    """
    completeness_info = {
        'is_complete': True,
        'missing_data': [],
        'category_counts': {},
        'total_expected': 0,
        'total_actual': 0
    }
    
    for category, result in category_results.items():
        actual_count = result['data_count']
        completeness_info['category_counts'][category] = {
            'expected': expected_count_per_category,
            'actual': actual_count,
            'missing': expected_count_per_category - actual_count
        }
        
        completeness_info['total_expected'] += expected_count_per_category
        completeness_info['total_actual'] += actual_count
        
        if actual_count < expected_count_per_category:
            completeness_info['is_complete'] = False
            completeness_info['missing_data'].append({
                'category': category,
                'expected': expected_count_per_category,
                'actual': actual_count,
                'missing': expected_count_per_category - actual_count
            })
    
    return completeness_info


def calculate_overall_aspect1_aspect2_from_categories(category_results: Dict) -> Tuple[float, float]:
    """
    Calculate Overall aspect1 and aspect2 percentages (based on weighted average of all categories)
    
    Args:
        category_results: Calculation results for each category
        
    Returns:
        (aspect1_percentage, aspect2_percentage) tuple
    """
    if not category_results:
        return 0.0, 0.0
    
    # Accumulate raw scores and denominators for each category
    total_aspect1_score = 0.0
    total_aspect1_denominator = 0.0
    total_aspect2_score = 0.0
    total_aspect2_denominator = 0.0
    
    for category, result in category_results.items():
        # Skip non-category data (like completeness, overall_score, etc.)
        if not isinstance(result, dict) or 'raw_scores' not in result:
            continue
            
        raw_scores = result['raw_scores']
        data_count = result['data_count']
        num_checklist = result['num_checklist']
        
        # aspect1 accumulation (use recorded denominator; compatible with old results)
        total_aspect1_score += raw_scores['aspect1_score']
        aspect1_denominator = raw_scores.get('aspect1_denominator', data_count)  # Compatible with old data
        total_aspect1_denominator += aspect1_denominator
        
        # aspect2 accumulation
        total_aspect2_score += raw_scores['aspect2_score']
        total_aspect2_denominator += num_checklist  # aspect2满分为total checklist count
    
    # Calculate percentage
    aspect1_percentage = (total_aspect1_score / total_aspect1_denominator * 100) if total_aspect1_denominator > 0 else 0.0
    aspect2_percentage = (total_aspect2_score / total_aspect2_denominator * 100) if total_aspect2_denominator > 0 else 0.0
    
    return aspect1_percentage, aspect2_percentage


def calculate_overall_score_from_categories(category_results: Dict) -> float:
    """
    Calculate Overall total score (based on weighted average of all categories)
    
    Args:
        category_results: Calculation results for each category
        
    Returns:
        Overall total score percentage
    """
    if not category_results:
        return 0.0
    
    total_raw_score = 0.0
    total_denominator = 0.0
    
    for category, result in category_results.items():
        raw_scores = result['raw_scores']
        data_count = result['data_count']
        num_checklist = result['num_checklist']
        
        # Accumulate raw scores and denominators
        total_raw_score += raw_scores['aspect1_score'] + raw_scores['aspect2_score']
        aspect1_denominator = raw_scores.get('aspect1_denominator', data_count)  # Compatible with old data
        total_denominator += num_checklist + aspect1_denominator
    
    # Calculate overall score
    overall_score = (total_raw_score / total_denominator * 100) if total_denominator > 0 else 0.0
    return overall_score


def process_hints_folder(hints_folder: str, expected_count: int = 50) -> Dict[str, Dict]:
    """
    Process all judged_{model_name}_{judge_model}.jsonl files in a single hints folder
    
    Args:
        hints_folder: hints folder path (如 infer_40_hints0)
        
    Returns:
        {model_name: {category: result_dict, 'overall': overall_score}} dictionary
    """
    if not os.path.exists(hints_folder):
        raise FileNotFoundError(f"Folder does not exist: {hints_folder}")
    
    # Find all matching files (support any judge model)
    pattern = os.path.join(hints_folder, "judged_*.jsonl")
    files = glob.glob(pattern)
    
    if not files:
        print(f"Warning: in folder {hints_folder} no files matching judged_*.jsonl format found")
        return {}
    
    print(f"\nProcessing folder: {hints_folder}")
    print(f"Found {len(files)} files:")
    for f in files:
        print(f"  - {os.path.basename(f)}")
    
    results = {}
    
    for file_path in files:
        filename = os.path.basename(file_path)
        # Use global model manager to extract model name
        model_name = model_manager.extract_model_name(filename)
        
        try:
            category_results = calculate_scores_for_file(file_path)
            overall_score = calculate_overall_score_from_categories(category_results)
            
            # Calculate Overall aspect1/aspect2 percentages (consistent with main experiment)
            overall_aspect1, overall_aspect2 = calculate_overall_aspect1_aspect2_from_categories(category_results)
            
            # Check data completeness
            completeness_info = check_data_completeness(category_results, expected_count)
            
            results[model_name] = category_results.copy()
            results[model_name]['overall'] = overall_score
            results[model_name]['overall_aspect1_percentage'] = overall_aspect1
            results[model_name]['overall_aspect2_percentage'] = overall_aspect2
            results[model_name]['completeness'] = completeness_info
            
            # Collect completeness issues to global list
            if not completeness_info['is_complete']:
                all_completeness_issues.append({
                    'folder': hints_folder,
                    'model': model_name,
                    'completeness_info': completeness_info,
                    'overall_score': overall_score,
                    'category_count': len(category_results)
                })
            
            # Only show processing completion info, don't show completeness issues (leave for final display)
            status_icon = "✓" if completeness_info['is_complete'] else "⚠️"
            print(f"{status_icon} {model_name}: 处理完成，Found {len(category_results)} categories, total score: {overall_score:.1f}%")
            
        except Exception as e:
            print(f"✗ {model_name}: Processing failed - {e}")
            continue
    
    return results


# Removed automatic discovery feature, now using fixed folder definitions


def process_all_hints_experiments(base_path: str, expected_count: int = 50) -> Dict[str, Dict[str, Dict]]:
    """
    Process all hints experiment folders
    
    Args:
        base_path: Base path containing all hints folders
        expected_count: 每个category期望的data count
        
    Returns:
        {hints_folder: {model_name: results}} nested dictionary
    """
    # Fixed defined hints folders
    hints_folders = [
        'judge_infer_50_hints0',  # No hint
        'judge_infer_50_hints1',  # Hint1
        'judge_infer_50_hints2',  # Hint2
        'judge_infer_50_hints3',  # Hint3
        'judge_infer_50_hints4'   # Hint1+Hint2+Hint3
    ]
    
    print(f"Processing specified hints experiment folders: {hints_folders}")
    
    all_results = {}
    
    for hints_folder in hints_folders:
        folder_path = os.path.join(base_path, hints_folder)
        
        if os.path.exists(folder_path):
            try:
                folder_results = process_hints_folder(folder_path, expected_count)
                if folder_results:
                    all_results[hints_folder] = folder_results
                else:
                    print(f"警告: {hints_folder} folder has no valid data")
            except Exception as e:
                print(f"Error: processing {hints_folder} failed - {e}")
        else:
            print(f"Warning: file夹 {folder_path} does not exist, skipping")
    
    return all_results


def process_main_experiment(base_path: str, expected_count: int = 50) -> Dict[str, Dict]:
    """
    处理主Experiment数据（自动寻找hints0folder的数据，按category显示）
    
    Args:
        base_path: 包含Experimentfolder的基础路径
        expected_count: 每个category期望的data count
        
    Returns:
        {model_name: {category: result_dict}} dictionary
    """
    # 使用固定的hints0folder作为主Experiment数据源
    hints0_folder = os.path.join(base_path, "judge_infer_50_hints0")
    
    if not os.path.exists(hints0_folder):
        print(f"警告: 主Experiment数据源folder {hints0_folder} 不存在")
        return {}
    
    # Find all matching files (support any judge model)
    pattern = os.path.join(hints0_folder, "judged_*.jsonl")
    files = glob.glob(pattern)
    
    if not files:
        print(f"Warning: in folder {hints0_folder} 中未Found主Experiment文件")
        return {}
    
    print(f"\n📊 Processing main experiment data (based on hints0 data):")
    print(f"Found {len(files)} main experiment files:")
    for f in files:
        print(f"  - {os.path.basename(f)}")
    
    results = {}
    
    for file_path in files:
        filename = os.path.basename(file_path)
        # Use global model manager to extract model name
        model_name = model_manager.extract_model_name(filename)
        
        try:
            category_results = calculate_scores_for_file(file_path)
            
            # Check data completeness
            completeness_info = check_data_completeness(category_results, expected_count)
            
            # Calculate Overall score
            overall_score = calculate_overall_score_from_categories(category_results)
            
            # Calculate Overall aspect1/aspect2 percentages
            overall_aspect1, overall_aspect2 = calculate_overall_aspect1_aspect2_from_categories(category_results)
            
            results[model_name] = category_results.copy()
            results[model_name]['completeness'] = completeness_info
            results[model_name]['overall_score'] = overall_score
            results[model_name]['overall_aspect1_percentage'] = overall_aspect1
            results[model_name]['overall_aspect2_percentage'] = overall_aspect2
            
            # Collect completeness issues to global list
            if not completeness_info['is_complete']:
                all_completeness_issues.append({
                    'folder': 'main_experiment',
                    'model': model_name,
                    'completeness_info': completeness_info,
                    'overall_score': overall_score,
                    'category_count': len(category_results)
                })
            
            # Only show processing completion info
            status_icon = "✓" if completeness_info['is_complete'] else "⚠️"
            print(f"{status_icon} {model_name}: 处理完成，Found {len(category_results)} categories, Overall: {overall_score:.1f}%")
            
        except Exception as e:
            print(f"✗ {model_name}: Processing failed - {e}")
            continue
    
    return results


def map_category_to_column(category: str) -> str:
    """
    Map category to Excel column name
    """
    category_mapping = {
        'philosophy': 'Phi',
        'Computer Science': 'CS', 
        'Law': 'Law',
        'economics': 'Econ',
        # Can add more mappings as needed
    }
    return category_mapping.get(category, category)


def create_comprehensive_excel(all_results: Dict, output_file: str):
    """
    Create Excel file containing both main and ablation experiment sheets
    
    Args:
        all_results: All results data for main and ablation experiments
        output_file: Output Excel file path
    """
    # Automatically create directory (if not exists)
    output_dir = os.path.dirname(output_file)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)
        print(f"Create directory: {output_dir}")
    
    # Separate main and ablation experiment data
    ablation_results = {}
    main_results = {}
    
    # 定义消融Experiment的folder名称
    ablation_folders = {
        'judge_infer_50_hints0',
        'judge_infer_50_hints1', 
        'judge_infer_50_hints2',
        'judge_infer_50_hints3',
        'judge_infer_50_hints4'
    }
    
    for key, value in all_results.items():
        if key in ablation_folders:
            ablation_results[key] = value
        else:
            # Main experiment data (model name as key)
            main_results[key] = value
    
    # Create workbook
    wb = Workbook()
    
    # Remove default Sheet
    wb.remove(wb.active)
    
    # === Create main experiment sheet ===
    main_ws = wb.create_sheet("main")
    
    if main_results:
        # Collect all appearing categories and sort
        all_categories = set()
        for model_results in main_results.values():
            # Only include real categories, exclude special fields
            for key, value in model_results.items():
                # Ensure this is a real category result, not a special field, and key is not None
                if (key is not None and 
                    key not in ['completeness', 'overall_score', 'overall_aspect1_percentage', 'overall_aspect2_percentage'] and
                    isinstance(value, dict) and 'calculated_scores' in value):
                    all_categories.add(key)
        
        # Map category to column name and sort
        column_mapping = {cat: map_category_to_column(cat) for cat in all_categories}
        # Filter out None values to prevent sorting errors
        valid_columns = [col for col in column_mapping.values() if col is not None]
        columns = ['Overall'] + sorted(set(valid_columns))
        
        print(f"Categories found in main experiment: {sorted(all_categories)}")
        print(f"Main experiment Excel columns: {columns}")
        
        # Set headers
        headers = ['Model'] + columns
        for col_idx, header in enumerate(headers, 1):
            cell = main_ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Use model manager to automatically sort models
        available_model_names = set(main_results.keys())
        available_models = model_manager.get_sorted_models(available_model_names)
        
        print(f"Models specified for main experiment (in order): {[display_name for _, display_name in available_models]}")
        
        # Fill data
        row_idx = 2
        for file_name, display_name in available_models:
            model_results = main_results[file_name]
            
            # Model name
            main_ws.cell(row=row_idx, column=1, value=display_name)
            
            # Overallcolumn: display aspect1/aspect2 format
            if 'overall_aspect1_percentage' in model_results and 'overall_aspect2_percentage' in model_results:
                aspect1_pct = model_results['overall_aspect1_percentage']
                aspect2_pct = model_results['overall_aspect2_percentage']
                main_ws.cell(row=row_idx, column=2, value=f"{aspect1_pct:.1f}/{aspect2_pct:.1f}")
            else:
                main_ws.cell(row=row_idx, column=2, value="")
            
            # Fill data for each category
            for col_idx, col_name in enumerate(columns[1:], 3):  # Start from column 3 (skip Overall)
                # Find corresponding category
                matching_category = None
                for category, mapped_name in column_mapping.items():
                    if mapped_name == col_name:
                        matching_category = category
                        break
                
                if matching_category and matching_category in model_results:
                    result = model_results[matching_category]
                    # Ensure this is a real category result with formatted_score field
                    if isinstance(result, dict) and 'formatted_score' in result:
                        formatted_score = result['formatted_score']
                        main_ws.cell(row=row_idx, column=col_idx, value=formatted_score)
                    else:
                        main_ws.cell(row=row_idx, column=col_idx, value="")
                else:
                    main_ws.cell(row=row_idx, column=col_idx, value="")
            
            row_idx += 1
        
        # Set column width
        for col in main_ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            main_ws.column_dimensions[column].width = adjusted_width
    
    # === Create sheets with same structure as main experiment for each hints condition ===
    if ablation_results:
        # Dynamic naming function
        def sheet_title_for_folder(folder_name: str) -> str:
            if 'hints0' in folder_name:
                return 'No hint'
            elif 'hints1' in folder_name:
                return 'Hint1'
            elif 'hints2' in folder_name:
                return 'Hint2'
            elif 'hints3' in folder_name:
                return 'Hint3'
            elif 'hints4' in folder_name:
                return 'Hint1+Hint2+Hint3'
            else:
                return folder_name

        for hints_folder in sorted(ablation_results.keys()):
            folder_results = ablation_results[hints_folder]
            title = sheet_title_for_folder(hints_folder)
            ws = wb.create_sheet(title)

            if folder_results:
                # Collect all appearing categories
                all_categories = set()
                for model_results in folder_results.values():
                    for key, value in model_results.items():
                        if (
                            key is not None and
                            key not in ['completeness', 'overall', 'overall_aspect1_percentage', 'overall_aspect2_percentage'] and
                            isinstance(value, dict) and 'calculated_scores' in value
                        ):
                            all_categories.add(key)

                column_mapping = {cat: map_category_to_column(cat) for cat in all_categories}
                valid_columns = [col for col in column_mapping.values() if col is not None]
                columns = ['Overall'] + sorted(set(valid_columns))

                # Headers
                headers = ['Model'] + columns
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')

                # Use model manager to automatically sort models
                available_model_names = set(folder_results.keys())
                available_models = model_manager.get_sorted_models(available_model_names)

                # Fill data
                row_idx = 2
                for file_name, display_name in available_models:
                    model_results = folder_results[file_name]
                    ws.cell(row=row_idx, column=1, value=display_name)

                    # Overall列：aspect1/aspect2
                    aspect1_pct = model_results.get('overall_aspect1_percentage', 0)
                    aspect2_pct = model_results.get('overall_aspect2_percentage', 0)
                    ws.cell(row=row_idx, column=2, value=f"{aspect1_pct:.1f}/{aspect2_pct:.1f}")

                    # Each category
                    for col_idx, col_name in enumerate(columns[1:], 3):
                        matching_category = None
                        for category, mapped_name in column_mapping.items():
                            if mapped_name == col_name:
                                matching_category = category
                                break
                        if matching_category and matching_category in model_results:
                            result = model_results[matching_category]
                            if isinstance(result, dict) and 'formatted_score' in result:
                                ws.cell(row=row_idx, column=col_idx, value=result['formatted_score'])
                            else:
                                ws.cell(row=row_idx, column=col_idx, value="")
                        else:
                            ws.cell(row=row_idx, column=col_idx, value="")

                    row_idx += 1

                # Auto-adjust column width
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 20)
                    ws.column_dimensions[column].width = adjusted_width

    # === Create ablation experiment sheet ===
    ablation_ws = wb.create_sheet("ablation experiment")
    
    if ablation_results:
        # 收集所有Model name
        all_models = set()
        for folder_results in ablation_results.values():
            for model_name, model_data in folder_results.items():
                # Ensure model data is valid (exclude special fields)
                if isinstance(model_data, dict) and any(key not in ['overall', 'completeness'] for key in model_data.keys()):
                    all_models.add(model_name)
        
        # Use model manager to automatically sort models
        available_models = model_manager.get_sorted_models(all_models)
        
        # Dynamic column name mapping (adapt to different hints numbers)
        def get_column_name(folder_name):
            if 'hints0' in folder_name:
                return 'No hint'
            elif 'hints1' in folder_name:
                return 'Hint1'
            elif 'hints2' in folder_name:
                return 'Hint2'
            elif 'hints3' in folder_name:
                return 'Hint3'
            elif 'hints4' in folder_name:
                return 'Hint1+Hint2+Hint3'
            else:
                return folder_name
        
        column_mapping = {folder: get_column_name(folder) for folder in ablation_results.keys()}
        
        # Set headers
        headers = ['ablation experiment'] + [column_mapping.get(folder, folder) for folder in sorted(ablation_results.keys())]
        for col_idx, header in enumerate(headers, 1):
            cell = ablation_ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        print(f"Ablation experiment Excel columns: {headers}")
        print(f"Models specified for ablation experiment (in order): {[display_name for _, display_name in available_models]}")
        
        # Fill data
        row_idx = 2
        for file_name, display_name in available_models:
            # Model name（显示名称）
            ablation_ws.cell(row=row_idx, column=1, value=display_name)
            
            # Fill data for each hints condition
            col_idx = 2
            for hints_folder in sorted(ablation_results.keys()):
                if file_name in ablation_results[hints_folder]:
                    # Use aspect1/aspect2 format (consistent with main experiment)
                    model_data = ablation_results[hints_folder][file_name]
                    aspect1_pct = model_data.get('overall_aspect1_percentage', 0)
                    aspect2_pct = model_data.get('overall_aspect2_percentage', 0)
                    ablation_ws.cell(row=row_idx, column=col_idx, value=f"{aspect1_pct:.1f}/{aspect2_pct:.1f}")
                else:
                    ablation_ws.cell(row=row_idx, column=col_idx, value="")
                col_idx += 1
            
            row_idx += 1
        
        # Set column width
        for col in ablation_ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ablation_ws.column_dimensions[column].width = adjusted_width
    
    # Save file
    wb.save(output_file)
    print(f"\nResults saved to: {output_file}")
    print(f"Included sheets: {[ws.title for ws in wb.worksheets]}")


def print_ablation_results(all_results: Dict[str, Dict[str, Dict]]):
    """
    Print ablation experiment results summary
    """
    print(f"\n📊 Ablation experiment results summary:")
    
    # Collect all models
    all_models = set()
    for folder_results in all_results.values():
        all_models.update(folder_results.keys())
    
    # Use model manager to automatically sort models
    available_models = model_manager.get_sorted_models(all_models)
    
    print(f"Processed {len(available_models)} specified models in {len(all_results)} experimental conditions")
    
    # Display results in specified order
    for file_name, display_name in available_models:
        print(f"\n🎯 {display_name}:")
        for hints_folder, folder_results in all_results.items():
            if file_name in folder_results:
                model_data = folder_results[file_name]
                # Use aspect1/aspect2 format (consistent with main experiment)
                aspect1_pct = model_data.get('overall_aspect1_percentage', 0)
                aspect2_pct = model_data.get('overall_aspect2_percentage', 0)
                # Dynamically get condition name
                if 'hints0' in hints_folder:
                    condition_name = 'No hint'
                elif 'hints1' in hints_folder:
                    condition_name = 'Hint1'
                elif 'hints2' in hints_folder:
                    condition_name = 'Hint2'
                elif 'hints3' in hints_folder:
                    condition_name = 'Hint3'
                elif 'hints4' in hints_folder:
                    condition_name = 'Hint1+Hint2+Hint3'
                else:
                    condition_name = hints_folder
                print(f"  📌 {condition_name}: {aspect1_pct:.1f}/{aspect2_pct:.1f}")


def print_completeness_issues():
    """
    Display all completeness issues
    """
    if not all_completeness_issues:
        print("\n✅ All data completeness checks passed, no missing data found")
        return
    
    print(f"\n⚠️  Found {len(all_completeness_issues)} data completeness issues:")
    print("="*60)
    
    for issue in all_completeness_issues:
        folder = issue['folder']
        model = issue['model']
        completeness_info = issue['completeness_info']
        overall_score = issue['overall_score']
        category_count = issue['category_count']
        
        print(f"\n🔍 {folder} - {model}:")
        print(f"   Total score: {overall_score:.1f}% | Categories: {category_count}")
        print(f"   Data incomplete: expected {completeness_info['total_expected']} entries, actual {completeness_info['total_actual']} 条")
        
        for missing in completeness_info['missing_data']:
            print(f"   - {missing['category']}: missing {missing['missing']} entries ({missing['actual']}/{missing['expected']})")


def save_completeness_issues_to_file(output_dir: str = "results"):
    """
    Save completeness issues to file
    
    Args:
        output_dir: Output directory
    """
    if not all_completeness_issues:
        return
    
    # 确保Output directory存在
    os.makedirs(output_dir, exist_ok=True)
    
    # Generate filename (with timestamp)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"data_completeness_issues_{timestamp}.txt"
    filepath = os.path.join(output_dir, filename)
    
    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(f"Data completeness issues report\n")
        f.write(f"Generation time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Found问题数量: {len(all_completeness_issues)}\n")
        f.write("="*60 + "\n\n")
        
        for issue in all_completeness_issues:
            folder = issue['folder']
            model = issue['model']
            completeness_info = issue['completeness_info']
            overall_score = issue['overall_score']
            category_count = issue['category_count']
            
            f.write(f"Experiment: {folder}\n")
            f.write(f"Model: {model}\n")
            f.write(f"Total score: {overall_score:.1f}%\n")
            f.write(f"Categories数量: {category_count}\n")
            f.write(f"Data completeness: 预期 {completeness_info['total_expected']} entries, actual {completeness_info['total_actual']} 条\n")
            f.write(f"Missing details:\n")
            
            for missing in completeness_info['missing_data']:
                f.write(f"  - {missing['category']}: missing {missing['missing']} entries ({missing['actual']}/{missing['expected']})\n")
            
            f.write("\n" + "-"*40 + "\n\n")
    
    print(f"\n📄 Completeness issues report saved to: {filepath}")


def print_main_results(main_results: Dict[str, Dict]):
    """
    打印主Experiment结果摘要
    """
    print(f"\n📊 主Experiment结果摘要:")
    print(f"Processed {len(main_results)} 个Model的数据")
    
    # Use model manager to automatically sort models
    available_model_names = set(main_results.keys())
    available_models = model_manager.get_sorted_models(available_model_names)
    
    # Display results in sorted order
    for file_name, display_name in available_models:
        print(f"\n🎯 {display_name}:")
        model_results = main_results[file_name]
        # Filter out real category data, exclude special fields
        for category, result in model_results.items():
            if category not in ['completeness', 'overall_score', 'overall_aspect1_percentage', 'overall_aspect2_percentage']:
                if isinstance(result, dict) and 'calculated_scores' in result:
                    calc = result['calculated_scores']
                    print(f"  📌 {category}: {calc['aspect1_total']*100:.1f}%/{calc['aspect2_total']*100:.1f}%")


def main():
    parser = argparse.ArgumentParser(description="综合Experiment分析：同时处理主Experiment和消融Experiment数据，生成包含两个sheet的Excel文件")
    parser.add_argument('base_path', help='包含Experiment数据的基础路径（包含hintsfolder和主Experiment文件）')
    parser.add_argument('--output', '-o', help='Output Excel file path（默认：results.xlsx）', default='results.xlsx')
    parser.add_argument('--verbose', '-v', action='store_true', help='Show verbose output')
    parser.add_argument('--expected-count', '-c', type=int, default=50, help='每个category期望的data count（默认：50）')
    
    args = parser.parse_args()
    
    # Clear global variables (prevent accumulation during multiple runs)
    global all_completeness_issues
    all_completeness_issues = []
    
    try:
        print("🚀 开始处理综合Experiment数据...")
        
        # 处理消融Experiment数据
        print("\n" + "="*50)
        print("🔬 消融Experiment数据处理")
        print("="*50)
        print(f"每个category期望data count: {args.expected_count}")
        ablation_results = process_all_hints_experiments(args.base_path, args.expected_count)
        
        # 处理主Experiment数据
        print("\n" + "="*50)
        print("📊 主Experiment数据处理")
        print("="*50)
        main_experiment_results = process_main_experiment(args.base_path, args.expected_count)
        
        # Check if data exists
        if not ablation_results and not main_experiment_results:
            print("❌ 没有成功处理任何Experiment数据")
            return 1
        
        # Print results summary
        if args.verbose:
            if ablation_results:
                print_ablation_results(ablation_results)
            if main_experiment_results:
                print_main_results(main_experiment_results)
        
        # Ensure output file has .xlsx extension
        output_file = args.output
        if not output_file.endswith('.xlsx'):
            output_file += '.xlsx'
        
        # 创建综合Excel文件
        print("\n" + "="*50)
        print("📝 Generate Excel report")
        print("="*50)
        
        # Merge all results for Excel generation
        all_results_for_excel = {}
        if ablation_results:
            all_results_for_excel.update(ablation_results)
        if main_experiment_results:
            all_results_for_excel.update(main_experiment_results)
        
        create_comprehensive_excel(all_results_for_excel, output_file)
        
        # Display and save completeness issues
        print("\n" + "="*50)
        print("🔍 Data completeness检查结果")
        print("="*50)
        print_completeness_issues()
        
        # Save completeness issues to file
        if all_completeness_issues:
            output_dir = os.path.dirname(output_file) or "results"
            save_completeness_issues_to_file(output_dir)
        
        # Statistics results
        ablation_count = len(ablation_results) if ablation_results else 0
        main_experiment_count = len(main_experiment_results) if main_experiment_results else 0
        
        print(f"\n✅ 综合Experiment数据处理完成！")
        print(f"🔬 消融Experiment: {ablation_count} 个Experiment条件")
        print(f"📊 主Experiment: {main_experiment_count} 个Model")
        print(f"📁 输出文件: {output_file}")
        if all_completeness_issues:
            print(f"⚠️  Found {len(all_completeness_issues)} data completeness issues，Please see above report for details")
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return 1
    
    return 0


if __name__ == "__main__":
    exit(main()) 